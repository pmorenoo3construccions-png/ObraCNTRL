#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generar_dashboard.py — OBRACTRL · O3 Construccions
====================================================
Llegeix els Excel de control de costos de cada obra,
extreu els valors actuals i actualitza el bloc JSON
de dades dins d'index.html.

Executa'l cada divendres (o quan vulguis) via publicar.bat.
Requereix: pip install openpyxl xlrd
"""

import json
import re
import os
import sys
from datetime import date, datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: Instal·la openpyxl: pip install openpyxl xlrd")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────
# CONFIGURACIÓ DE RUTES
# ─────────────────────────────────────────────────────────────
# BASE = carpeta on es troba aquest script (repo GitHub)
BASE = os.path.dirname(os.path.abspath(__file__))

# OBRAS_BASE = carpeta local sincronitzada amb Google Drive
# on es troben les carpetes de cada obra i els seus Excel
OBRAS_BASE = r"C:\Users\rokit\GOOGLE DRIVE\OBRAS PEDRO"

PATHS = {
    "rambla_costos": os.path.join(
        OBRAS_BASE, "OBRAS ACTIVAS", "MT_CAP_RAMBLA",
        "04_VARIS", "CONTROL_COSTOS",
        "CONTROL COSTOS CAP Rambla.xlsx"
    ),
    "xirgu_costos": os.path.join(
        OBRAS_BASE, "OBRAS ACTIVAS", "MARGARIDA_XIRGU_24",
        "04_VARIS", "Control costos",
        "Control costos Margarida Xirgu 24 Sitges.xlsx"
    ),
    "xirgu_cert_dir": os.path.join(
        OBRAS_BASE, "OBRAS ACTIVAS", "MARGARIDA_XIRGU_24",
        "02_CERTIFICACIO_PRODUCCIO"
    ),
    "irla_costos": os.path.join(
        OBRAS_BASE, "OBRAS ACTIVAS", "JOSEP_IRLA_32",
        "04_VARIS", "Control costos",
        "Control costes Josep Irla Sitges.xlsx"
    ),
    "irla_cert_dir": os.path.join(
        OBRAS_BASE, "OBRAS ACTIVAS", "JOSEP_IRLA_32",
        "02_CERTIFICACIO_PRODUCCIO"
    ),
    "index_html": os.path.join(BASE, "index.html"),
}


# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────
def fmt_eur(val, decimals=0):
    """Formata un número com a import en €. Ex: 103061.12 → '103.061 €'"""
    if val is None or val == 0:
        return "0 €"
    sign = "+" if val > 0 else ""
    rounded = round(float(val), decimals)
    if decimals == 0:
        rounded = int(rounded)
        s = f"{abs(rounded):,}".replace(",", ".")
    else:
        s = f"{abs(rounded):,.{decimals}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{sign if val >= 0 else '-'}{s} €".strip()


def fmt_k(val):
    """Formata en milers. Ex: 1096000 → '1.096K€'"""
    k = round(val / 1000)
    return f"{k:,}€".replace(",", ".").replace("€", "K€")


def fmt_pct(val):
    """Formata percentatge. Ex: 23.612 → '23,6%'"""
    return f"{val:.1f}%".replace(".", ",")


def safe_float(cell):
    """Retorna el valor numèric d'una cel·la o None."""
    if cell is None:
        return None
    v = cell.value if hasattr(cell, 'value') else cell
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    return None


def find_numeric_rows(ws, col=None, min_val=1000):
    """Troba files on hi ha valors numèrics significatius."""
    results = []
    for row in ws.iter_rows():
        for cell in row:
            if col is not None and cell.column != col:
                continue
            v = safe_float(cell)
            if v and abs(v) >= min_val:
                results.append((cell.row, cell.column, v))
    return results


# ─────────────────────────────────────────────────────────────
# EXTRACCIÓ: CAP RAMBLA
# ─────────────────────────────────────────────────────────────
def extract_rambla():
    """Extreu cost real acumulat i nombre de certificacions de Cap Rambla."""
    path = PATHS["rambla_costos"]
    data = {
        "cost_real": 103061.12,
        "n_certs": 6,
        "ultima_cert": "feb26",
    }

    if not os.path.exists(path):
        print(f"  [RAMBLA] Fitxer no trobat: {path}")
        return data

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active

        # Buscar el total (fila TOTAL o RESUM) — valors > 80.000
        totals = []
        for row in ws.iter_rows():
            for cell in row:
                v = safe_float(cell)
                if v and 80000 <= v <= 500000:
                    totals.append(v)

        if totals:
            data["cost_real"] = max(totals)

        # Comptar certificacions (cerquem paraules 'cert' en una fila)
        n_certs = 0
        for row in ws.iter_rows():
            for cell in row:
                val = str(cell.value or "").lower()
                if "certif" in val or "cert." in val:
                    n_certs += 1
        if n_certs > 0:
            data["n_certs"] = min(n_certs, 15)  # màx raonable

        print(f"  [RAMBLA] Cost real: {data['cost_real']:,.0f} € · Certs: {data['n_certs']}")
    except Exception as e:
        print(f"  [RAMBLA] Error llegint Excel: {e} — usant valors per defecte")

    return data


# ─────────────────────────────────────────────────────────────
# EXTRACCIÓ: MARGARIDA XIRGU
# ─────────────────────────────────────────────────────────────
def extract_xirgu():
    """Extreu cost real i facturat (última certificació) de Xirgu."""
    data = {
        "cost_real": 69928.0,
        "facturat": 91554.0,
        "n_cert": 5,
        "ultima_cert_prev": 66726.0,
        "ultima_cert_val": 24828.0,
    }

    # Cost real del Excel de control de costos
    path_cost = PATHS["xirgu_costos"]
    if os.path.exists(path_cost):
        try:
            wb = openpyxl.load_workbook(path_cost, data_only=True)
            ws = wb.active
            totals = []
            for row in ws.iter_rows():
                for cell in row:
                    v = safe_float(cell)
                    if v and 20000 <= v <= 300000:
                        totals.append(v)
            if totals:
                data["cost_real"] = max(totals)
            print(f"  [XIRGU] Cost real: {data['cost_real']:,.0f} €")
        except Exception as e:
            print(f"  [XIRGU] Error costos: {e}")
    else:
        print(f"  [XIRGU] Fitxer costos no trobat, usant valors per defecte")

    # Última certificació: buscar l'Excel o PDF de cert més recent
    cert_dir = PATHS["xirgu_cert_dir"]
    if os.path.exists(cert_dir):
        # Buscar fitxers de certificació ordenats per data de modificació
        cert_files = []
        for f in os.listdir(cert_dir):
            fl = f.lower()
            if ("certif" in fl or "cert" in fl) and (fl.endswith(".xls") or fl.endswith(".xlsx")):
                full = os.path.join(cert_dir, f)
                cert_files.append((os.path.getmtime(full), full, f))
        cert_files.sort(reverse=True)

        for _, full_path, fname in cert_files:
            try:
                wb2 = openpyxl.load_workbook(full_path, data_only=True)
                ws2 = wb2.active
                # Buscar valor ORIGEN (acumulat)
                origen_vals = []
                for row in ws2.iter_rows():
                    for cell in row:
                        lbl = str(cell.value or "").lower()
                        if "origen" in lbl or "total" in lbl or "acum" in lbl:
                            # Buscar valor numèric a la mateixa fila
                            for c2 in ws2[cell.row]:
                                v = safe_float(c2)
                                if v and 50000 <= v <= 500000:
                                    origen_vals.append(v)
                if origen_vals:
                    data["facturat"] = max(origen_vals)
                    print(f"  [XIRGU] Facturat (origen): {data['facturat']:,.0f} € — {fname}")
                    break
            except Exception as e:
                print(f"  [XIRGU] Error cert {fname}: {e}")

    # Calcular número de cert a partir del nom del fitxer
    if cert_files:
        import re as re2
        last_name = cert_files[0][2]
        m = re2.search(r'n[ºo]?\s*(\d+)', last_name, re2.IGNORECASE)
        if m:
            data["n_cert"] = int(m.group(1))

    return data


# ─────────────────────────────────────────────────────────────
# EXTRACCIÓ: JOSEP IRLA 32
# ─────────────────────────────────────────────────────────────
def extract_irla():
    """Extreu cost real i estat de certificació de Irla/Sitges."""
    data = {
        "cost_real": 921475.0,
        "cert_acum": 1078506.19,
        "pressupost": 1160970.18,
        "cert_mes": 101089.0,
        "n_cert": 15,
        "ultima_cert": "gen26",
    }

    path_cost = PATHS["irla_costos"]
    if os.path.exists(path_cost):
        try:
            wb = openpyxl.load_workbook(path_cost, data_only=True)
            ws = wb.active
            totals = []
            for row in ws.iter_rows():
                for cell in row:
                    v = safe_float(cell)
                    if v and 800000 <= v <= 2000000:
                        totals.append(v)
            if totals:
                data["cost_real"] = max(totals)
            print(f"  [IRLA] Cost real: {data['cost_real']:,.0f} €")
        except Exception as e:
            print(f"  [IRLA] Error costos: {e}")
    else:
        print(f"  [IRLA] Fitxer costos no trobat, usant valors per defecte")

    # Última certificació
    cert_dir = PATHS["irla_cert_dir"]
    if os.path.exists(cert_dir):
        cert_files = []
        for root, dirs, files in os.walk(cert_dir):
            for f in files:
                fl = f.lower()
                if ("certif" in fl) and (fl.endswith(".xls") or fl.endswith(".xlsx")):
                    full = os.path.join(root, f)
                    cert_files.append((os.path.getmtime(full), full, f))
        cert_files.sort(reverse=True)

        for _, full_path, fname in cert_files[:3]:
            try:
                wb2 = openpyxl.load_workbook(full_path, data_only=True)
                for sheet in wb2.sheetnames:
                    if "resum" in sheet.lower() or "resumen" in sheet.lower():
                        ws2 = wb2[sheet]
                        for row in ws2.iter_rows():
                            for cell in row:
                                lbl = str(cell.value or "").upper()
                                if "ORIGEN" in lbl or "ACUMULAD" in lbl:
                                    for c2 in ws2[cell.row]:
                                        v = safe_float(c2)
                                        if v and 900000 <= v <= 2000000:
                                            data["cert_acum"] = v
                                            print(f"  [IRLA] Cert acum: {v:,.0f} € — {fname}")
                break
            except Exception as e:
                print(f"  [IRLA] Error cert {fname}: {e}")

    return data


# ─────────────────────────────────────────────────────────────
# CONSTRUCCIÓ DEL BLOC JSON
# ─────────────────────────────────────────────────────────────
def build_json(rambla, xirgu, irla):
    today_str = date.today().strftime("%d/%m/%Y")

    # Xirgu calculats
    xirgu_marge = xirgu["facturat"] - xirgu["cost_real"]
    xirgu_marge_pct = (xirgu_marge / xirgu["facturat"] * 100) if xirgu["facturat"] > 0 else 0
    xirgu_cost_bar_w = round(xirgu["cost_real"] / xirgu["facturat"] * 100) if xirgu["facturat"] > 0 else 0
    xirgu_marge_bar_w = 100 - xirgu_cost_bar_w
    n_cert = xirgu["n_cert"]
    cert_ordinal = {1:"1ª",2:"2ª",3:"3ª",4:"4ª",5:"5ª",6:"6ª",7:"7ª",8:"8ª"}.get(n_cert, f"{n_cert}ª")

    # Irla calculats
    irla_pendent = irla["pressupost"] - irla["cert_acum"]
    irla_pct = (irla["cert_acum"] / irla["pressupost"] * 100) if irla["pressupost"] > 0 else 0
    irla_pct_bar = round(irla_pct, 1)

    # KPI totals (HAD = 0 perquè no ha començat)
    kpi_cost = rambla["cost_real"] + xirgu["cost_real"] + irla["cost_real"]
    kpi_cert = xirgu["facturat"] + irla["cert_acum"]

    # Número de cert Irla en text
    irla_n = irla["n_cert"]
    mes_map = {1:"gen",2:"feb",3:"mar",4:"abr",5:"mai",6:"jun",7:"jul",8:"ago",9:"set",10:"oct",11:"nov",12:"des"}
    irla_mes_label = irla.get("ultima_cert", "gen26")

    data = {
        "kpi": {
            "cost_total": fmt_k(kpi_cost),
            "cert_total": fmt_k(kpi_cert),
            "cert_sub": f"Irla (cert. {irla_n}ª · {irla_mes_label})"
        },
        "rambla": {
            "cost_real_eur": fmt_eur(rambla["cost_real"]),
            "mini_total": fmt_eur(rambla["cost_real"]),
            "panel_sub": f"{fmt_eur(rambla['cost_real'])} · {rambla['n_certs']} certifs."
        },
        "xirgu": {
            "fact_label": f"Facturat acum. (cert. {cert_ordinal})",
            "facturat_eur": fmt_eur(xirgu["facturat"]),
            "cost_real_sub": f"Cost real: {fmt_eur(xirgu['cost_real'])}",
            "marge_sub": f"Marge: {fmt_eur(xirgu_marge)} ({fmt_pct(xirgu_marge_pct)})",
            "prog_fact": fmt_eur(xirgu["facturat"]).replace(" ", ""),
            "prog_cost": fmt_eur(xirgu["cost_real"]).replace(" ", ""),
            "prog_marge": f"+{fmt_pct(xirgu_marge_pct)}" if xirgu_marge >= 0 else fmt_pct(xirgu_marge_pct),
            "cost_bar_w": xirgu_cost_bar_w,
            "marge_bar_w": xirgu_marge_bar_w,
            "desv_val": f"+{fmt_pct(xirgu_marge_pct)}" if xirgu_marge >= 0 else fmt_pct(xirgu_marge_pct),
            "desv_text": (
                f"Facturat {fmt_eur(xirgu['facturat'])} − Cost {fmt_eur(xirgu['cost_real'])}<br>"
                f"= Marge {fmt_eur(xirgu_marge)} · Pressupost obert"
            ),
            "mini_fact": fmt_eur(xirgu["facturat"]),
            "mini_cost": fmt_eur(xirgu["cost_real"]),
            "mini_marge": fmt_eur(xirgu_marge),
            "alert_sub": (
                f"Facturat acum. cert. {cert_ordinal}: {fmt_eur(xirgu['facturat'])} · "
                f"Cost real: {fmt_eur(xirgu['cost_real'])} · Marge: {fmt_eur(xirgu_marge)} "
                f"({fmt_pct(xirgu_marge_pct)}). Pressupost obert: el client afegeix feines sobre la marxa."
            ),
            "panel_sub": (
                f"{fmt_eur(xirgu['cost_real'])} real / {fmt_eur(xirgu['facturat'])} facturat (pressupost obert)"
            ),
            "panel_pct": f"+{fmt_pct(xirgu_marge_pct)}" if xirgu_marge >= 0 else fmt_pct(xirgu_marge_pct)
        },
        "irla": {
            "cert_acum_sub": f"Cert. acum.: {fmt_eur(irla['cert_acum'])}",
            "pend_sub": f"Pendent: {fmt_eur(irla_pendent)}",
            "mini_cert_acum": fmt_eur(irla["cert_acum"]),
            "mini_cert_mes": fmt_eur(irla["cert_mes"]),
            "mini_pendent": fmt_eur(irla_pendent),
            "mini_pct": fmt_pct(irla_pct),
            "cert_pct_bar": irla_pct_bar,
            "desv_text": (
                f"Cost real ~{fmt_k(irla['cost_real'])} vs cert. {fmt_k(irla['cert_acum'])}<br>"
                f"Marge positiu · Recta final"
            ),
            "panel_sub": f"~{fmt_k(irla['cost_real'])} cost / {fmt_k(irla['cert_acum'])} cert."
        }
    }
    return data


# ─────────────────────────────────────────────────────────────
# ACTUALITZAR index.html
# ─────────────────────────────────────────────────────────────
def update_html(data_dict):
    path = PATHS["index_html"]
    with open(path, "r", encoding="utf-8") as f:
        html = f.read()

    new_json = json.dumps(data_dict, ensure_ascii=False, indent=2)

    # Substituir el bloc JSON entre les etiquetes del <script id="obraData">
    pattern = r'(<script id="obraData" type="application/json">\s*\n)(\{.*?\})(\s*\n</script>)'
    replacement = r'\g<1>' + new_json + r'\g<3>'
    new_html, count = re.subn(pattern, replacement, html, flags=re.DOTALL)

    if count == 0:
        print("ERROR: No s'ha trobat el bloc <script id=\"obraData\">. Comprova index.html.")
        return False

    with open(path, "w", encoding="utf-8") as f:
        f.write(new_html)

    print(f"  index.html actualitzat ✓ — {count} substitució(ns)")
    return True


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 60)
    print(f"OBRACTRL generar_dashboard.py — {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("=" * 60)

    print("\n[1/4] Extraient dades Cap Rambla...")
    rambla = extract_rambla()

    print("\n[2/4] Extraient dades Margarida Xirgu...")
    xirgu = extract_xirgu()

    print("\n[3/4] Extraient dades Irla / Sitges...")
    irla = extract_irla()

    print("\n[4/4] Actualitzant index.html...")
    data = build_json(rambla, xirgu, irla)
    ok = update_html(data)

    if ok:
        print("\n✅ Dashboard generat correctament.")
        print(f"   Rambla: {rambla['cost_real']:,.0f} €")
        print(f"   Xirgu:  {xirgu['facturat']:,.0f} € facturat / {xirgu['cost_real']:,.0f} € cost")
        print(f"   Irla:   {irla['cert_acum']:,.0f} € cert acum.")
    else:
        print("\n❌ Error actualitzant el dashboard.")
        sys.exit(1)
